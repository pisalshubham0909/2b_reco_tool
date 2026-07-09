import json
import re
import pandas as pd
import numpy as np
import gc
from openpyxl import load_workbook

# Shadow built-in float in module scope to handle currency symbols, commas, and malformed inputs robustly
_builtin_float = float
def float(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, _builtin_float)):
        return _builtin_float(val)
    try:
        val_str = str(val).replace(',', '').replace('₹', '').replace('$', '').strip()
        if not val_str or val_str.lower() in ('none', 'nan', 'null', ''):
            return 0.0
        return _builtin_float(val_str)
    except Exception:
        return 0.0

def get_json_section(data, key):

    """
    Search recursively or through common GSTR-2B JSON wrappers for a specific key.
    """
    if not isinstance(data, dict):
        return []
    
    raw_val = []
    # Direct match
    if key in data:
        raw_val = data[key]
    # Check standard 'data' wrappers
    elif 'data' in data:
        inner = data['data']
        if isinstance(inner, dict):
            if key in inner:
                raw_val = inner[key]
            elif 'data' in inner:
                double_inner = inner['data']
                if isinstance(double_inner, dict) and key in double_inner:
                    raw_val = double_inner[key]
                
    if not raw_val:
        # Exhaustive search for the key (as a fallback)
        for k, v in data.items():
            if k == key:
                raw_val = v
                break
            if isinstance(v, dict):
                res = get_json_section(v, key)
                if res:
                    raw_val = res
                    break

    if isinstance(raw_val, list):
        return [item for item in raw_val if isinstance(item, dict)]
    return []

def clean_invoice_number(inv_no):
    """
    Cleans invoice number for robust matching:
    - Standardizes to uppercase.
    - Strips whitespace.
    - Removes non-alphanumeric characters.
    - Strips leading zeros.
    - Strips float trailing decimals (.0 / .00).
    """
    if pd.isna(inv_no) or inv_no is None:
        return ""
    inv_str = str(inv_no).strip().upper()
    # Remove floating point suffix if parsed as float
    if inv_str.endswith('.0'):
        inv_str = inv_str[:-2]
    elif inv_str.endswith('.00'):
        inv_str = inv_str[:-3]
    # Remove all non-alphanumeric characters (e.g., slash, dash, spaces)
    cleaned = re.sub(r'[^A-Z0-9]', '', inv_str)
    # Strip leading zeros
    cleaned = cleaned.lstrip('0')
    return cleaned

def parse_date(date_val):
    """
    Parses various date formats to pd.Timestamp.
    """
    if pd.isna(date_val) or date_val is None:
        return pd.NaT
    if isinstance(date_val, pd.Timestamp):
        return date_val
    if isinstance(date_val, (int, _builtin_float)):
        # Excel numeric date handling
        try:
            return pd.to_datetime(date_val, unit='D', origin='1899-12-30')
        except:
            return pd.NaT
            
    date_str = str(date_val).strip()
    for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d', '%d-%b-%Y', '%d %b %Y'):
        try:
            return pd.to_datetime(date_str, format=fmt)
        except ValueError:
            continue
    try:
        return pd.to_datetime(date_str)
    except:
        return pd.NaT

def parse_gstr2b_json(json_content_or_path, file_name="GSTR2B.json"):
    """
    Parses a GSTR-2B JSON file and flattens it into a pandas DataFrame.
    Supports B2B, B2BA, CDNR, CDNRA, ISD, ISDA, IMPG, and IMPGSEZ sections.
    """
    if isinstance(json_content_or_path, str):
        try:
            with open(json_content_or_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except Exception as e:
            # Try load as raw string if path fails
            data = json.loads(json_content_or_path)
    else:
        # File-like object
        data = json.load(json_content_or_path)

    # Handle double serialization (where JSON contains a stringified JSON object)
    if isinstance(data, str):
        try:
            data = json.loads(data)
        except Exception:
            pass

    if not isinstance(data, dict):
        raise ValueError("Invalid GSTR-2B JSON structure: Root element must be a dictionary/object.")

    # Helper function to recursively lowercase all dictionary keys to resolve portal casing discrepancies
    def lowercase_keys_recursive(d):
        if isinstance(d, list):
            return [lowercase_keys_recursive(v) for v in d]
        if isinstance(d, dict):
            return {k.lower(): lowercase_keys_recursive(v) for k, v in d.items()}
        return d

    data = lowercase_keys_recursive(data)

    # Attempt to extract return period and recipient GSTIN
    rtn_period = data.get('rtnprd') or data.get('fp')
    recipient_gstin = data.get('gstin') or data.get('rcptgstin')
    
    if not rtn_period and 'data' in data and isinstance(data['data'], dict):
        rtn_period = data['data'].get('rtnprd') or data['data'].get('fp')
        recipient_gstin = data['data'].get('gstin') or data['data'].get('rcptgstin')

    documents = []

    # Helper function to extract document list from standard formats
    def get_doc_list(parent_obj):
        if not isinstance(parent_obj, dict):
            return []
        doc_list = []
        if 'doclist' in parent_obj:
            doc_list = parent_obj['doclist']
        elif 'inv' in parent_obj:
            doc_list = parent_obj['inv']
        elif 'nt' in parent_obj:
            doc_list = parent_obj['nt']
        elif 'boe' in parent_obj:
            doc_list = parent_obj['boe']
            
        if isinstance(doc_list, list):
            return [item for item in doc_list if isinstance(item, dict)]
        return []

    # Helper function to extract tax values from items robustly, checking 'itm_det' and synonyms
    def extract_tax_from_items(items):
        txval_sum = 0.0
        igst_sum = 0.0
        cgst_sum = 0.0
        sgst_sum = 0.0
        cess_sum = 0.0
        if not items:
            return txval_sum, igst_sum, cgst_sum, sgst_sum, cess_sum
            
        for item in items:
            if not isinstance(item, dict):
                continue
            det = item.get('itm_det') or item.get('itmdet')
            if not isinstance(det, dict):
                det = item
                
            txval_sum += float(det.get('txval') or 0.0)
            igst_sum += float(det.get('iamt') or det.get('igst') or 0.0)
            cgst_sum += float(det.get('camt') or det.get('cgst') or 0.0)
            sgst_sum += float(det.get('samt') or det.get('sgst') or 0.0)
            cess_sum += float(det.get('csamt') or det.get('cess') or 0.0)
            
        return txval_sum, igst_sum, cgst_sum, sgst_sum, cess_sum

    # 1. Parse B2B Invoices
    b2b_list = get_json_section(data, 'b2b')
    for supplier in b2b_list:
        ctin = supplier.get('ctin', '').strip().upper()
        cname = supplier.get('lglnm') or supplier.get('trdnm') or ""
        doc_list = get_doc_list(supplier)
        
        for inv in doc_list:
            inum = inv.get('inum', '').strip()
            idt = inv.get('idt') or inv.get('dt')
            val = float(inv.get('val', 0.0))
            itcelg = inv.get('itcelg', 'Y').strip().upper()
            pos = inv.get('pos') or ""
            rchrg = inv.get('rchrg') or inv.get('rc') or 'N'
            filing_date = inv.get('flddt') or inv.get('fld_dt') or inv.get('filing_date') or ""
            gstr3b_status = inv.get('g3bfil') or inv.get('g3bfilingstatus') or inv.get('g3bstatus') or 'N'
            
            items = inv.get('items') or inv.get('itms') or []
            txval, igst, cgst, sgst, cess = extract_tax_from_items(items)
            
            documents.append({
                'supplier_gstin': ctin,
                'supplier_name': cname,
                'doc_num': inum,
                'clean_doc_num': clean_invoice_number(inum),
                'doc_date': parse_date(idt),
                'doc_type': 'INV',
                'taxable_val': round(txval, 2),
                'igst': round(igst, 2),
                'cgst': round(cgst, 2),
                'sgst': round(sgst, 2),
                'cess': round(cess, 2),
                'total_val': round(val, 2),
                'pos': pos,
                'rchrg': 'Yes' if rchrg in ('Y', 'YES') else 'No',
                'itc_eligibility': 'Eligible' if itcelg in ('Y', 'YES') else 'Ineligible',
                'filing_date': parse_date(filing_date),
                'gstr3b_status': 'Yes' if gstr3b_status in ('Y', 'YES', 'FILING_STATUS_FILED') else 'No',
                'section': 'B2B Invoices',
                'is_amended': False,
                'original_doc_num': None,
                'original_doc_date': None,
                'rtn_period': rtn_period,
                'source': 'GSTR-2B',
                'source_file': file_name
            })

    # 2. Parse B2BA (Amended B2B Invoices)
    b2ba_list = get_json_section(data, 'b2ba')
    for supplier in b2ba_list:
        ctin = supplier.get('ctin', '').strip().upper()
        cname = supplier.get('lglnm') or supplier.get('trdnm') or ""
        doc_list = get_doc_list(supplier)
        
        for inv in doc_list:
            inum = inv.get('inum', '').strip()
            idt = inv.get('idt') or inv.get('dt')
            oinum = inv.get('oinum', '').strip()
            oidt = inv.get('oidt')
            val = float(inv.get('val', 0.0))
            itcelg = inv.get('itcelg', 'Y').strip().upper()
            pos = inv.get('pos') or ""
            rchrg = inv.get('rchrg') or inv.get('rc') or 'N'
            filing_date = inv.get('flddt') or inv.get('fld_dt') or inv.get('filing_date') or ""
            gstr3b_status = inv.get('g3bfil') or inv.get('g3bfilingstatus') or inv.get('g3bstatus') or 'N'
            
            items = inv.get('items') or inv.get('itms') or []
            txval, igst, cgst, sgst, cess = extract_tax_from_items(items)
            
            documents.append({
                'supplier_gstin': ctin,
                'supplier_name': cname,
                'doc_num': inum,
                'clean_doc_num': clean_invoice_number(inum),
                'doc_date': parse_date(idt),
                'doc_type': 'INV',
                'taxable_val': round(txval, 2),
                'igst': round(igst, 2),
                'cgst': round(cgst, 2),
                'sgst': round(sgst, 2),
                'cess': round(cess, 2),
                'total_val': round(val, 2),
                'pos': pos,
                'rchrg': 'Yes' if rchrg in ('Y', 'YES') else 'No',
                'itc_eligibility': 'Eligible' if itcelg in ('Y', 'YES') else 'Ineligible',
                'filing_date': parse_date(filing_date),
                'gstr3b_status': 'Yes' if gstr3b_status in ('Y', 'YES', 'FILING_STATUS_FILED') else 'No',
                'section': 'B2B Amendments',
                'is_amended': True,
                'original_doc_num': oinum,
                'original_doc_date': parse_date(oidt),
                'rtn_period': rtn_period,
                'source': 'GSTR-2B',
                'source_file': file_name
            })

    # 3. Parse CDNR (Credit / Debit Notes)
    cdnr_list = get_json_section(data, 'cdnr')
    for supplier in cdnr_list:
        ctin = supplier.get('ctin', '').strip().upper()
        cname = supplier.get('lglnm') or supplier.get('trdnm') or ""
        doc_list = get_doc_list(supplier)
        
        for nt in doc_list:
            nt_num = (nt.get('nt_num') or nt.get('ntnum') or '').strip()
            nt_dt = nt.get('nt_dt') or nt.get('ntdt') or nt.get('dt')
            nt_ty = nt.get('nt_ty', '').strip().upper() # C = Credit Note, D = Debit Note
            val = float(nt.get('val', 0.0))
            itcelg = nt.get('itcelg', 'Y').strip().upper()
            pos = nt.get('pos') or ""
            rchrg = nt.get('rchrg') or nt.get('rc') or 'N'
            filing_date = nt.get('flddt') or nt.get('fld_dt') or nt.get('filing_date') or ""
            gstr3b_status = nt.get('g3bfil') or nt.get('g3bfilingstatus') or nt.get('g3bstatus') or 'N'
            
            items = nt.get('items') or nt.get('itms') or []
            txval, igst, cgst, sgst, cess = extract_tax_from_items(items)
            
            # Credit notes represent negative values (ITC reduction)
            sign = -1.0 if nt_ty == 'C' else 1.0
            doc_type = 'CRN' if nt_ty == 'C' else 'DBN'
            
            documents.append({
                'supplier_gstin': ctin,
                'supplier_name': cname,
                'doc_num': nt_num,
                'clean_doc_num': clean_invoice_number(nt_num),
                'doc_date': parse_date(nt_dt),
                'doc_type': doc_type,
                'taxable_val': round(txval * sign, 2),
                'igst': round(igst * sign, 2),
                'cgst': round(cgst * sign, 2),
                'sgst': round(sgst * sign, 2),
                'cess': round(cess * sign, 2),
                'total_val': round(val * sign, 2),
                'pos': pos,
                'rchrg': 'Yes' if rchrg in ('Y', 'YES') else 'No',
                'itc_eligibility': 'Eligible' if itcelg in ('Y', 'YES') else 'Ineligible',
                'filing_date': parse_date(filing_date),
                'gstr3b_status': 'Yes' if gstr3b_status in ('Y', 'YES', 'FILING_STATUS_FILED') else 'No',
                'section': 'Credit/Debit Notes',
                'is_amended': False,
                'original_doc_num': None,
                'original_doc_date': None,
                'rtn_period': rtn_period,
                'source': 'GSTR-2B',
                'source_file': file_name
            })

    # 4. Parse CDNRA (Amended Credit / Debit Notes)
    cdnra_list = get_json_section(data, 'cdnra')
    for supplier in cdnra_list:
        ctin = supplier.get('ctin', '').strip().upper()
        cname = supplier.get('lglnm') or supplier.get('trdnm') or ""
        doc_list = get_doc_list(supplier)
        
        for nt in doc_list:
            nt_num = (nt.get('nt_num') or nt.get('ntnum') or '').strip()
            nt_dt = nt.get('nt_dt') or nt.get('ntdt') or nt.get('dt')
            ont_num = (nt.get('ont_num') or nt.get('ontnum') or '').strip()
            ont_dt = nt.get('ont_dt') or nt.get('ontdt')
            nt_ty = nt.get('nt_ty', '').strip().upper()
            val = float(nt.get('val', 0.0))
            itcelg = nt.get('itcelg', 'Y').strip().upper()
            pos = nt.get('pos') or ""
            rchrg = nt.get('rchrg') or nt.get('rc') or 'N'
            filing_date = nt.get('flddt') or nt.get('fld_dt') or nt.get('filing_date') or ""
            gstr3b_status = nt.get('g3bfil') or nt.get('g3bfilingstatus') or nt.get('g3bstatus') or 'N'
            
            items = nt.get('items') or nt.get('itms') or []
            txval, igst, cgst, sgst, cess = extract_tax_from_items(items)
            
            sign = -1.0 if nt_ty == 'C' else 1.0
            doc_type = 'CRN' if nt_ty == 'C' else 'DBN'
            
            documents.append({
                'supplier_gstin': ctin,
                'supplier_name': cname,
                'doc_num': nt_num,
                'clean_doc_num': clean_invoice_number(nt_num),
                'doc_date': parse_date(nt_dt),
                'doc_type': doc_type,
                'taxable_val': round(txval * sign, 2),
                'igst': round(igst * sign, 2),
                'cgst': round(cgst * sign, 2),
                'sgst': round(sgst * sign, 2),
                'cess': round(cess * sign, 2),
                'total_val': round(val * sign, 2),
                'pos': pos,
                'rchrg': 'Yes' if rchrg in ('Y', 'YES') else 'No',
                'itc_eligibility': 'Eligible' if itcelg in ('Y', 'YES') else 'Ineligible',
                'filing_date': parse_date(filing_date),
                'gstr3b_status': 'Yes' if gstr3b_status in ('Y', 'YES', 'FILING_STATUS_FILED') else 'No',
                'section': 'Credit/Debit Notes Amendments',
                'is_amended': True,
                'original_doc_num': ont_num,
                'original_doc_date': parse_date(ont_dt),
                'rtn_period': rtn_period,
                'source': 'GSTR-2B',
                'source_file': file_name
            })

    # 5. Parse ISD Invoices
    isd_list = get_json_section(data, 'isd')
    for supplier in isd_list:
        ctin = supplier.get('ctin', '').strip().upper()
        cname = supplier.get('lglnm') or supplier.get('trdnm') or ""
        doc_list = get_doc_list(supplier)
        
        for inv in doc_list:
            docnum = (inv.get('docnum') or inv.get('doc_num') or inv.get('inum') or '').strip()
            docdt = inv.get('docdt') or inv.get('doc_dt') or inv.get('idt')
            val = float(inv.get('val', 0.0))
            itcelg = inv.get('itcelg', 'Y').strip().upper()
            pos = inv.get('pos') or ""
            filing_date = inv.get('flddt') or inv.get('fld_dt') or inv.get('filing_date') or ""
            gstr3b_status = inv.get('g3bfil') or inv.get('g3bfilingstatus') or inv.get('g3bstatus') or 'Y'
            
            # ISD distribution values may be split rate-wise, sum them up
            items = inv.get('items') or inv.get('itms') or []
            txval, igst, cgst, sgst, cess = extract_tax_from_items(items)
            
            # Fallback if items are missing
            if not items:
                txval = float(inv.get('txval') or inv.get('tx_val') or val)
                igst = float(inv.get('igst') or inv.get('iamt') or 0.0)
                cgst = float(inv.get('cgst') or inv.get('camt') or 0.0)
                sgst = float(inv.get('sgst') or inv.get('samt') or 0.0)
                cess = float(inv.get('cess') or inv.get('csamt') or 0.0)

            documents.append({
                'supplier_gstin': ctin,
                'supplier_name': cname,
                'doc_num': docnum,
                'clean_doc_num': clean_invoice_number(docnum),
                'doc_date': parse_date(docdt),
                'doc_type': 'ISD',
                'taxable_val': round(txval, 2),
                'igst': round(igst, 2),
                'cgst': round(cgst, 2),
                'sgst': round(sgst, 2),
                'cess': round(cess, 2),
                'total_val': round(val, 2),
                'pos': pos,
                'rchrg': 'No', # ISD distribute input credits, no reverse charge
                'itc_eligibility': 'Eligible' if itcelg in ('Y', 'YES') else 'Ineligible',
                'filing_date': parse_date(filing_date),
                'gstr3b_status': 'Yes' if gstr3b_status in ('Y', 'YES') else 'No',
                'section': 'ISD Invoices',
                'is_amended': False,
                'original_doc_num': None,
                'original_doc_date': None,
                'rtn_period': rtn_period,
                'source': 'GSTR-2B',
                'source_file': file_name
            })

    # 6. Parse ISDA Invoices (ISD Amendments)
    isda_list = get_json_section(data, 'isda')
    for supplier in isda_list:
        ctin = supplier.get('ctin', '').strip().upper()
        cname = supplier.get('lglnm') or supplier.get('trdnm') or ""
        doc_list = get_doc_list(supplier)
        
        for inv in doc_list:
            docnum = (inv.get('docnum') or inv.get('doc_num') or inv.get('inum') or '').strip()
            docdt = inv.get('docdt') or inv.get('doc_dt') or inv.get('idt')
            odocnum = (inv.get('odocnum') or inv.get('odoc_num') or inv.get('oinum') or '').strip()
            odocdt = inv.get('odocdt') or inv.get('odoc_dt') or inv.get('oidt')
            val = float(inv.get('val', 0.0))
            itcelg = inv.get('itcelg', 'Y').strip().upper()
            pos = inv.get('pos') or ""
            filing_date = inv.get('flddt') or inv.get('fld_dt') or inv.get('filing_date') or ""
            gstr3b_status = inv.get('g3bfil') or inv.get('g3bfilingstatus') or inv.get('g3bstatus') or 'Y'
            
            items = inv.get('items') or inv.get('itms') or []
            txval, igst, cgst, sgst, cess = extract_tax_from_items(items)
            
            if not items:
                txval = float(inv.get('txval') or inv.get('tx_val') or val)
                igst = float(inv.get('igst') or inv.get('iamt') or 0.0)
                cgst = float(inv.get('cgst') or inv.get('camt') or 0.0)
                sgst = float(inv.get('sgst') or inv.get('samt') or 0.0)
                cess = float(inv.get('cess') or inv.get('csamt') or 0.0)

            documents.append({
                'supplier_gstin': ctin,
                'supplier_name': cname,
                'doc_num': docnum,
                'clean_doc_num': clean_invoice_number(docnum),
                'doc_date': parse_date(docdt),
                'doc_type': 'ISD',
                'taxable_val': round(txval, 2),
                'igst': round(igst, 2),
                'cgst': round(cgst, 2),
                'sgst': round(sgst, 2),
                'cess': round(cess, 2),
                'total_val': round(val, 2),
                'pos': pos,
                'rchrg': 'No',
                'itc_eligibility': 'Eligible' if itcelg in ('Y', 'YES') else 'Ineligible',
                'filing_date': parse_date(filing_date),
                'gstr3b_status': 'Yes' if gstr3b_status in ('Y', 'YES') else 'No',
                'section': 'ISD Amendments',
                'is_amended': True,
                'original_doc_num': odocnum,
                'original_doc_date': parse_date(odocdt),
                'rtn_period': rtn_period,
                'source': 'GSTR-2B',
                'source_file': file_name
            })

    # 7. Parse IMPG (Import of Goods)
    impg_list = get_json_section(data, 'impg')
    for boe in impg_list:
        boe_num = str(boe.get('boe_num') or boe.get('boenum') or boe.get('boenm') or '').strip()
        boe_dt = boe.get('boe_dt') or boe.get('boedt')
        val = float(boe.get('boe_val') or boe.get('val', 0.0))
        txval = float(boe.get('txval') or boe.get('tx_val') or 0.0)
        igst = float(boe.get('igst') or boe.get('iamt') or 0.0)
        cess = float(boe.get('cess') or boe.get('csamt') or 0.0)
        itcelg = boe.get('itcelg', 'Y').strip().upper()
        port_cd = boe.get('port_cd') or boe.get('port_code') or 'CUSTOMS'
        
        documents.append({
            'supplier_gstin': 'IMPORT',
            'supplier_name': f'Import from Port: {port_cd}',
            'doc_num': boe_num,
            'clean_doc_num': clean_invoice_number(boe_num),
            'doc_date': parse_date(boe_dt),
            'doc_type': 'IMPG',
            'taxable_val': round(txval, 2),
            'igst': round(igst, 2),
            'cgst': 0.0,
            'sgst': 0.0,
            'cess': round(cess, 2),
            'total_val': round(val, 2),
            'pos': '97', # Outside India State Code
            'rchrg': 'No',
            'itc_eligibility': 'Eligible' if itcelg in ('Y', 'YES') else 'Ineligible',
            'filing_date': pd.NaT,
            'gstr3b_status': 'Yes',
            'section': 'Import of Goods',
            'is_amended': False,
            'original_doc_num': None,
            'original_doc_date': None,
            'rtn_period': rtn_period,
            'source': 'GSTR-2B',
            'source_file': file_name
        })

    # 8. Parse IMPGSEZ (Import from SEZ units)
    impgsez_list = get_json_section(data, 'impgsez')
    for boe in impgsez_list:
        boe_num = str(boe.get('boe_num') or boe.get('boenum') or boe.get('boenm') or '').strip()
        boe_dt = boe.get('boe_dt') or boe.get('boedt')
        val = float(boe.get('boe_val') or boe.get('val', 0.0))
        txval = float(boe.get('txval') or boe.get('tx_val') or 0.0)
        igst = float(boe.get('igst') or boe.get('iamt') or 0.0)
        cess = float(boe.get('cess') or boe.get('csamt') or 0.0)
        itcelg = boe.get('itcelg', 'Y').strip().upper()
        
        # SEZ imports usually have the actual SEZ supplier's GSTIN
        ctin = (boe.get('ctin') or boe.get('gstin') or 'SEZ-IMPORT').strip().upper()
        cname = boe.get('lglnm') or boe.get('trdnm') or "SEZ Supplier"
        pos = boe.get('pos') or ""

        documents.append({
            'supplier_gstin': ctin,
            'supplier_name': cname,
            'doc_num': boe_num,
            'clean_doc_num': clean_invoice_number(boe_num),
            'doc_date': parse_date(boe_dt),
            'doc_type': 'IMPG',
            'taxable_val': round(txval, 2),
            'igst': round(igst, 2),
            'cgst': 0.0,
            'sgst': 0.0,
            'cess': round(cess, 2),
            'total_val': round(val, 2),
            'pos': pos,
            'rchrg': 'No',
            'itc_eligibility': 'Eligible' if itcelg in ('Y', 'YES') else 'Ineligible',
            'filing_date': pd.NaT,
            'gstr3b_status': 'Yes',
            'section': 'Import from SEZ',
            'is_amended': False,
            'original_doc_num': None,
            'original_doc_date': None,
            'rtn_period': rtn_period,
            'source': 'GSTR-2B',
            'source_file': file_name
        })

    # Clean data & free memory
    del data
    gc.collect()

    # Create dataframe
    if not documents:
        return pd.DataFrame(columns=[
            'supplier_gstin', 'supplier_name', 'doc_num', 'clean_doc_num', 
            'doc_date', 'doc_type', 'taxable_val', 'igst', 'cgst', 'sgst', 
            'cess', 'total_val', 'pos', 'rchrg', 'itc_eligibility', 
            'filing_date', 'gstr3b_status', 'section', 'is_amended', 
            'original_doc_num', 'original_doc_date', 'rtn_period', 'source', 'source_file'
        ])
        
    df = pd.DataFrame(documents)
    return df

def parse_large_excel_streaming(file_path_or_buffer, sheet_name=None):
    """
    Streaming row-by-row Excel parser to read large spreadsheets (up to 1GB)
    with low memory overhead.
    """
    wb = load_workbook(file_path_or_buffer, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    
    rows = ws.iter_rows(values_only=True)
    
    # Read headers
    try:
        headers_raw = next(rows)
    except StopIteration:
        wb.close()
        return pd.DataFrame()
        
    headers = []
    # Fill in None or duplicate headers dynamically
    for idx, h in enumerate(headers_raw):
        if h is None:
            headers.append(f"Column_{idx}")
        else:
            h_str = str(h).strip()
            if h_str in headers:
                headers.append(f"{h_str}_{idx}")
            else:
                headers.append(h_str)

    data = []
    # Stream rows
    for r in rows:
        # Avoid loading extra empty rows at the end of Excel
        if all(val is None for val in r):
            continue
        # Zip row content to dict matching headers size
        row_dict = dict(zip(headers, r[:len(headers)]))
        data.append(row_dict)
        
    df_raw = pd.DataFrame(data)
    wb.close()
    
    # Force clean garbage collection
    del data
    gc.collect()
    
    return df_raw

def parse_purchase_register(file_path_or_buffer, col_mapping, sheet_name=None, credit_note_convention='auto'):
    """
    Parses the Purchase Register Excel/CSV file using the provided column mappings.
    Standardizes and normalizes the data for reconciliation.
    Uses memory-optimized streaming reader for Excel.
    """
    # 1. Load file
    if hasattr(file_path_or_buffer, 'columns') or type(file_path_or_buffer).__name__ == 'DataFrame':
        df_raw = file_path_or_buffer
    elif isinstance(file_path_or_buffer, str) and file_path_or_buffer.lower().endswith('.csv'):
        # For CSV, read chunked or with low_memory flag
        df_raw = pd.read_csv(file_path_or_buffer, low_memory=True)
    else:
        # Check if legacy .xls file
        is_xls = False
        if isinstance(file_path_or_buffer, str) and file_path_or_buffer.lower().endswith('.xls'):
            is_xls = True
        elif hasattr(file_path_or_buffer, 'name') and str(file_path_or_buffer.name).lower().endswith('.xls'):
            is_xls = True
            
        # Check file size (only use read_only streaming for files > 10MB to avoid openpyxl BytesIO zip bugs)
        is_large = False
        if hasattr(file_path_or_buffer, 'size') and file_path_or_buffer.size > 10 * 1024 * 1024:
            is_large = True
            
        if is_xls:
            df_raw = pd.read_excel(file_path_or_buffer, sheet_name=sheet_name)
        elif is_large:
            df_raw = parse_large_excel_streaming(file_path_or_buffer, sheet_name=sheet_name)
        else:
            df_raw = pd.read_excel(file_path_or_buffer, sheet_name=sheet_name)

    # 2. Extract and map fields
    df = pd.DataFrame()
    
    # Map Supplier GSTIN
    gst_col = col_mapping.get('supplier_gstin')
    if gst_col and gst_col in df_raw.columns:
        df['supplier_gstin'] = df_raw[gst_col].astype(str).str.strip().str.upper()
    else:
        raise ValueError("Supplier GSTIN column mapping is required.")
        
    # Map Document Number
    num_col = col_mapping.get('doc_num')
    if num_col and num_col in df_raw.columns:
        df['doc_num'] = df_raw[num_col].astype(str).str.strip()
        df['clean_doc_num'] = df['doc_num'].apply(clean_invoice_number)
    else:
        raise ValueError("Document/Invoice Number column mapping is required.")
        
    # Map Document Date
    date_col = col_mapping.get('doc_date')
    if date_col and date_col in df_raw.columns:
        df['doc_date'] = df_raw[date_col].apply(parse_date)
    else:
        raise ValueError("Invoice/Document Date column mapping is required.")
        
    # Map Supplier Name
    name_col = col_mapping.get('supplier_name')
    if name_col and name_col in df_raw.columns:
        df['supplier_name'] = df_raw[name_col].fillna("").astype(str).str.strip()
    else:
        df['supplier_name'] = ""
        
    # Map Document Type
    type_col = col_mapping.get('doc_type')
    if type_col and type_col in df_raw.columns:
        df['doc_type_raw'] = df_raw[type_col].fillna("").astype(str).str.strip().str.upper()
    else:
        df['doc_type_raw'] = 'INV'
        
    # Map Place of Supply
    pos_col = col_mapping.get('pos')
    if pos_col and pos_col in df_raw.columns:
        df['pos'] = df_raw[pos_col].fillna("").astype(str).str.strip()
    else:
        df['pos'] = ""

    # Map Reverse Charge
    rc_col = col_mapping.get('rchrg')
    if rc_col and rc_col in df_raw.columns:
        def parse_rc(val):
            val_str = str(val).strip().upper()
            return 'Yes' if val_str in ('Y', 'YES', 'TRUE', '1') else 'No'
        df['rchrg'] = df_raw[rc_col].apply(parse_rc)
    else:
        df['rchrg'] = 'No'

    # Map Monetary values
    tx_val_col = col_mapping.get('taxable_val')
    if not tx_val_col or tx_val_col not in df_raw.columns:
        raise ValueError("Taxable Value column mapping is required.")
        
    val_cols = {
        'taxable_val': tx_val_col,
        'igst': col_mapping.get('igst'),
        'cgst': col_mapping.get('cgst'),
        'sgst': col_mapping.get('sgst'),
        'cess': col_mapping.get('cess'),
        'total_val': col_mapping.get('total_val')
    }
    
    for key, col_name in val_cols.items():
        if col_name and col_name in df_raw.columns:
            df[key] = pd.to_numeric(df_raw[col_name].astype(str).replace(r'[\$,₹,]', '', regex=True), errors='coerce').fillna(0.0)
        else:
            df[key] = 0.0

    # Map PR Period
    period_col = col_mapping.get('pr_period')
    if period_col and period_col in df_raw.columns:
        df['pr_period'] = df_raw[period_col].fillna("").astype(str).str.strip()
    else:
        def fallback_period(row):
            dt = row.get('doc_date')
            if pd.notna(dt):
                return dt.strftime('%m%Y')
            return ""
        df['pr_period'] = df.apply(fallback_period, axis=1)

    # Determine standard document type (INV, CRN, DBN, IMPG, ISD)
    def determine_doc_type(row):
        t_raw = str(row.get('doc_type_raw', '')).upper()
        if 'CREDIT' in t_raw or 'CRN' in t_raw or 'CN' == t_raw or 'CDN' in t_raw:
            return 'CRN'
        elif 'DEBIT' in t_raw or 'DBN' in t_raw or 'DN' == t_raw:
            return 'DBN'
        elif 'IMPORT' in t_raw or 'IMPG' in t_raw or 'BOE' in t_raw or 'BILL OF ENTRY' in t_raw:
            return 'IMPG'
        elif 'ISD' in t_raw or 'ISDA' in t_raw or 'DISTRIB' in t_raw:
            return 'ISD'
        elif row.get('taxable_val') < 0:
            return 'CRN'
        return 'INV'

    df['doc_type'] = df.apply(determine_doc_type, axis=1)

    # Normalize Credit Note values to negative
    def normalize_signs(row):
        is_cn = (row['doc_type'] == 'CRN')
        is_negative = (row['taxable_val'] < 0)
        if is_cn and not is_negative:
            row['taxable_val'] = -abs(row['taxable_val'])
            row['igst'] = -abs(row['igst'])
            row['cgst'] = -abs(row['cgst'])
            row['sgst'] = -abs(row['sgst'])
            row['cess'] = -abs(row['cess'])
            row['total_val'] = -abs(row['total_val'])
        return row

    if credit_note_convention in ('auto', 'negative'):
        df = df.apply(normalize_signs, axis=1)

    # Additional standard fields for Books
    df['is_amended'] = False
    df['original_doc_num'] = None
    df['original_doc_date'] = None
    df['rtn_period'] = None
    df['itc_eligibility'] = 'Eligible'
    df['filing_date'] = pd.NaT
    df['gstr3b_status'] = 'Yes'
    df['section'] = 'Purchase Register'
    df['source'] = 'Books'
    df['source_file'] = 'Purchase Register'

    # Round all monetary values
    for val_col in ['taxable_val', 'igst', 'cgst', 'sgst', 'cess', 'total_val']:
        df[val_col] = df[val_col].round(2)

    # Clean up raw DataFrame references and collect garbage
    del df_raw
    gc.collect()

    return df

def auto_detect_columns(columns):
    """
    Auto-detect columns from a list of column headers using common keywords.
    """
    col_lower = [str(c).lower().strip() for c in columns]
    detected = {}
    
    keywords = {
        'supplier_gstin': ['gstin', 'gst', 'gst no', 'gstin/uin', 'supplier gst', 'ctin', 'party gst', 'vendor gstin'],
        'supplier_name': ['supplier name', 'vendor name', 'party name', 'name', 'supplier_name', 'party_name', 'lglname', 'legal name', 'vendor name'],
        'doc_num': ['invoice number', 'invoice no', 'inv no', 'bill no', 'voucher no', 'document number', 'doc no', 'invoice_no', 'inv_num', 'inum', 'document no', 'bill number', 'document no'],
        'doc_date': ['invoice date', 'date', 'inv date', 'bill date', 'voucher date', 'doc date', 'invoice_dt', 'idt', 'invoice_date', 'bill date', 'document date'],
        'taxable_val': ['taxable value', 'taxable amount', 'taxable amt', 'taxable val', 'assessable value', 'taxable_value', 'taxable_amt', 'txval', 'purchase value'],
        'igst': ['igst', 'integrated tax', 'igst amount', 'igst amt', 'igst_amt', 'igst_val'],
        'cgst': ['cgst', 'central tax', 'cgst amount', 'cgst amt', 'cgst_amt', 'cgst_val'],
        'sgst': ['sgst', 'state tax', 'sgst amount', 'sgst amt', 'sgst_amt', 'sgst_val', 'utgst', 'utgst amt', 'utgst_amount'],
        'cess': ['cess', 'cess amount', 'cess amt', 'cess_amt', 'cess_val', 'cec'],
        'total_val': ['total value', 'invoice value', 'total amount', 'inv value', 'bill amount', 'invoice_val', 'val', 'total_amt', 'gross value', 'invoice amount', 'document value', 'doc value'],
        'doc_type': ['document type', 'doc type', 'voucher type', 'vtype', 'type', 'doc_type', 'voucher name'],
        'pos': ['pos', 'place of supply', 'place_of_supply', 'state code', 'supply state'],
        'rchrg': ['rchrg', 'reverse charge', 'rcm', 'rc', 'reverse_charge'],
        'pr_period': ['period', 'month', 'return period', 'return_period', 'pr period', 'pr_period', 'rtnprd', 'month/year', 'reco period', 'reported period', 'reported_period']
    }
    
    for field, terms in keywords.items():
        matched = False
        for term in terms:
            for idx, col in enumerate(col_lower):
                if term == col or (len(term) >= 3 and term in col):
                    detected[field] = columns[idx]
                    matched = True
                    break
            if matched:
                break
                
    return detected

def parse_gstr2b_excel(file_path_or_obj):
    """
    Parses a GSTR-2B Excel spreadsheet directly downloaded from the GST Portal.
    Extracts records from B2B, B2BA, CDNR, CDNRA, ISD, IMPG, and IMPGSEZ sheets.
    """
    wb = load_workbook(file_path_or_obj, read_only=True, data_only=True)
    documents = []
    
    # Configuration for each worksheet type
    sheet_configs = {
        'b2b': {'doc_type': 'INV', 'section': 'B2B Invoices', 'is_cdnr': False},
        'b2ba': {'doc_type': 'INV', 'section': 'B2B Amendments', 'is_cdnr': False},
        'cdnr': {'section': 'Credit/Debit Notes', 'is_cdnr': True},
        'cdnra': {'section': 'Amended Credit/Debit Notes', 'is_cdnr': True},
        'isd': {'doc_type': 'ISD', 'section': 'ISD Invoices', 'is_cdnr': False},
        'impg': {'doc_type': 'IMPG', 'section': 'Import of Goods', 'is_cdnr': False},
        'impgsez': {'doc_type': 'IMPG', 'section': 'SEZ Import of Goods', 'is_cdnr': False}
    }
    
    # Case-insensitive worksheet names index map
    sheet_names_lower = {name.lower().strip(): name for name in wb.sheetnames}
    
    for key, config in sheet_configs.items():
        matched_sheet = None
        for name_lower, actual_name in sheet_names_lower.items():
            # Match config key as substring of sheet name case-insensitively, avoiding false matches
            if key == 'b2b':
                if 'b2b' in name_lower and 'b2ba' not in name_lower:
                    matched_sheet = actual_name
                    break
            elif key == 'cdnr':
                if ('cdnr' in name_lower or 'credit' in name_lower or 'debit' in name_lower or 'note' in name_lower) and 'cdnra' not in name_lower:
                    matched_sheet = actual_name
                    break
            elif key == 'impg':
                if ('impg' in name_lower or 'import' in name_lower) and 'impgsez' not in name_lower and 'sez' not in name_lower:
                    matched_sheet = actual_name
                    break
            else:
                if key in name_lower:
                    matched_sheet = actual_name
                    break
                    
        if matched_sheet:
            ws = wb[matched_sheet]
            
            # Find the header row by searching first 20 rows
            header_row_idx = None
            headers = []
            row_generator = ws.iter_rows(values_only=True)
            
            for r_idx, row in enumerate(row_generator, start=1):
                if r_idx > 20:
                    break
                row_str = [str(x).lower().strip() if x is not None else "" for x in row]
                # Look for header indicators
                if any('gstin' in s or 'ctin' in s or 'gst' in s or 'invoice' in s or 'doc' in s or 'bill' in s or 'boe' in s or 'note' in s for s in row_str):
                    header_row_idx = r_idx
                    headers = [str(x).strip() for x in row]
                    break
            
            if header_row_idx is None:
                continue
                
            # Iterate rows below header
            ws_rows = ws.iter_rows(values_only=True)
            for _ in range(header_row_idx):
                next(ws_rows, None)
                
            for row in ws_rows:
                # Skip empty or total rows
                if not row or row[0] is None or str(row[0]).strip().lower() in ('', 'none', 'total', 'grand total'):
                    continue
                    
                row_dict = {}
                for col_idx, col_name in enumerate(headers):
                    if col_idx < len(row) and col_name:
                        row_dict[col_name.lower().strip()] = row[col_idx]
                        
                # Extract fields using extremely flexible keyword/substring matches
                ctin_col = next((k for k in row_dict.keys() if 'gstin' in k or 'ctin' in k or 'gst' in k), None)
                ctin = str(row_dict.get(ctin_col or '')).strip().upper() if ctin_col else ''
                if not ctin or ctin.lower() in ('none', 'total', 'grand total'):
                    continue
                    
                name_col = next((k for k in row_dict.keys() if 'name' in k or 'trade' in k or 'legal' in k or 'supplier' in k), None)
                cname = str(row_dict.get(name_col or '')).strip() if name_col else ''
                
                inum_col = next((k for k in row_dict.keys() if 'invoice' in k or 'document' in k or 'doc num' in k or 'doc no' in k or 'note num' in k or 'note no' in k or 'boe num' in k or 'boe no' in k or k in ('inum', 'num', 'number', 'no')), None)
                inum = str(row_dict.get(inum_col or '')).strip() if inum_col else ''
                if not inum or inum.lower() in ('none', ''):
                    continue
                    
                idt_col = next((k for k in row_dict.keys() if 'date' in k or 'dt' in k), None)
                idt = row_dict.get(idt_col or '') if idt_col else None
                
                val_col = next((k for k in row_dict.keys() if 'value' in k or 'val' in k or 'amt' in k or 'amount' in k), None)
                val = float(row_dict.get(val_col or 0.0) or 0.0) if val_col else 0.0
                
                pos_col = next((k for k in row_dict.keys() if 'place' in k or 'pos' in k or 'supply' in k), None)
                pos = str(row_dict.get(pos_col or '')).strip() if pos_col else ''
                
                rc_col = next((k for k in row_dict.keys() if 'reverse' in k or 'rcm' in k or 'charge' in k or 'rc' in k), None)
                rchrg = str(row_dict.get(rc_col or 'N')).strip().upper() if rc_col else 'N'
                
                txval_col = next((k for k in row_dict.keys() if 'taxable' in k or 'txval' in k or 'tx_val' in k), None)
                txval = float(row_dict.get(txval_col or 0.0) or 0.0) if txval_col else 0.0
                
                igst_col = next((k for k in row_dict.keys() if 'integrated' in k or 'igst' in k or 'iamt' in k or 'int' in k), None)
                igst = float(row_dict.get(igst_col or 0.0) or 0.0) if igst_col else 0.0
                
                cgst_col = next((k for k in row_dict.keys() if 'central' in k or 'cgst' in k or 'camt' in k or 'cen' in k), None)
                cgst = float(row_dict.get(cgst_col or 0.0) or 0.0) if cgst_col else 0.0
                
                sgst_col = next((k for k in row_dict.keys() if 'sgst' in k or 'samt' in k or 'state/ut' in k or 'state tax' in k or 'utgst' in k), None)
                sgst = float(row_dict.get(sgst_col or 0.0) or 0.0) if sgst_col else 0.0
                
                cess_col = next((k for k in row_dict.keys() if 'cess' in k or 'csamt' in k), None)
                cess = float(row_dict.get(cess_col or 0.0) or 0.0) if cess_col else 0.0
                
                itc_col = next((k for k in row_dict.keys() if 'eligible' in k or 'eligibility' in k or 'availability' in k or 'itc' in k or 'avail' in k), None)
                itcelg = str(row_dict.get(itc_col or 'Y')).strip().upper() if itc_col else 'Y'
                itc_eligibility = 'Eligible' if 'ineligible' not in itcelg.lower() and itcelg in ('Y', 'YES', 'ELIGIBLE') else 'Ineligible'
                
                fld_col = next((k for k in row_dict.keys() if 'filing date' in k or 'filing dt' in k or 'flddt' in k or 'fld_dt' in k or 'filed date' in k), None)
                filing_date = row_dict.get(fld_col or '') if fld_col else ''
                
                g3b_col = next((k for k in row_dict.keys() if 'gstr-3b' in k or '3b' in k or 'filing status' in k or 'status' in k), None)
                gstr3b_status = 'Yes' if str(row_dict.get(g3b_col or 'N')).strip().upper() in ('Y', 'YES', 'FILED') else 'No'
                
                period_col = next((k for k in row_dict.keys() if 'period' in k or 'month' in k or 'year' in k or 'fp' in k or 'rtn' in k), None)
                rtn_period = str(row_dict.get(period_col or '')).strip() if period_col else ''
                
                doc_type = config.get('doc_type', 'INV')
                sign = 1.0
                
                if config['is_cdnr']:
                    ty_col = next((k for k in row_dict.keys() if 'type' in k or 'note type' in k or 'nt_ty' in k), None)
                    ty = str(row_dict.get(ty_col or '')).strip().upper() if ty_col else ''
                    if 'credit' in ty.lower() or 'c' in ty.lower():
                        doc_type = 'CRN'
                        sign = -1.0
                    else:
                        doc_type = 'DBN'
                        sign = 1.0
                        
                documents.append({
                    'supplier_gstin': ctin,
                    'supplier_name': cname,
                    'doc_num': inum,
                    'clean_doc_num': clean_invoice_number(inum),
                    'doc_date': parse_date(idt),
                    'doc_type': doc_type,
                    'taxable_val': round(txval * sign, 2),
                    'igst': round(igst * sign, 2),
                    'cgst': round(cgst * sign, 2),
                    'sgst': round(sgst * sign, 2),
                    'cess': round(cess * sign, 2),
                    'total_val': round(val * sign, 2),
                    'pos': pos,
                    'rchrg': 'Yes' if rchrg in ('Y', 'YES') else 'No',
                    'itc_eligibility': itc_eligibility,
                    'filing_date': parse_date(filing_date),
                    'gstr3b_status': gstr3b_status,
                    'section': config['section'],
                    'is_amended': 'amendment' in config['section'].lower(),
                    'original_doc_num': None,
                    'original_doc_date': None,
                    'rtn_period': rtn_period,
                    'source': 'GSTR-2B',
                    'source_file': getattr(file_path_or_obj, 'name', 'GSTR2B_Portal.xlsx')
                })
                
    wb.close()
    if not documents:
        return pd.DataFrame()
    return pd.DataFrame(documents)
